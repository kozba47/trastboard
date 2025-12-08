from datetime import date, datetime
import re
import base64

from flask import Flask, render_template, jsonify, request
from openpyxl import load_workbook

from config import EXCEL_FILE

app = Flask(__name__)


# ---------------------- ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ----------------------


def parse_excel_date(cell) -> date | None:
    """
    Универсальный парсер даты из ячейки Excel.

    Поддерживает:
    - datetime / date -> date
    - строки вида:
        '2025-12-08'
        '08.12.2025'
        '8.12.2025'
        '08.12.25'
        '08-12-2025'
        '08/12/2025'
        '08.12.2025 г.' и т.п.

    Любые хвосты типа 'г.' / 'год' отбрасываются.
    Если распарсить не удалось – возвращаем None.
    """
    # Уже Python-даты
    if isinstance(cell, datetime):
        return cell.date()
    if isinstance(cell, date):
        return cell

    # Строки
    if isinstance(cell, str):
        s = cell.strip()
        if not s:
            return None

        # убираем типичные хвосты
        s = (
            s.replace("года", "")
            .replace("год", "")
            .replace("г.", " ")
            .replace("г", " ")
        )
        s = s.strip()

        # оставляем только цифры, разделители и пробел/двоеточие
        s = re.sub(r"[^0-9.\-/: ]", "", s).strip()
        if not s:
            return None

        # Если есть время через пробел — берём отдельно весь s и отдельно часть до пробела
        if " " in s:
            first_part = s.split(" ")[0].strip()
        else:
            first_part = s

        candidates = [s]
        if first_part != s:
            candidates.append(first_part)

        # Набор форматов, которые пробуем
        formats = [
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%d.%m.%Y",
            "%d.%m.%y",
            "%d-%m-%Y",
            "%d-%m-%y",
            "%d/%m/%Y",
            "%d/%m/%y",
        ]

        for cand in candidates:
            for fmt in formats:
                try:
                    dt = datetime.strptime(cand, fmt)
                    # Поддержка двухзначного года: 25 -> 2025
                    if dt.year < 100:
                        dt = dt.replace(year=2000 + dt.year)
                    return dt.date()
                except ValueError:
                    continue

    # Остальное (числа, None и т.п.) не трогаем как даты
    return None


def get_header_metrics() -> dict:
    """
    Данные для шапки: курс USD и Brent из листа 'Курсы'.
    """
    usd_rate = None
    brent_price = None

    try:
        if not EXCEL_FILE.exists():
            return {"usd_rate": None, "brent_price": None}

        wb = load_workbook(EXCEL_FILE, data_only=True)
        if "Курсы" not in wb.sheetnames:
            return {"usd_rate": None, "brent_price": None}

        ws = wb["Курсы"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return {"usd_rate": None, "brent_price": None}

        header = rows[0]
        data_rows = rows[1:]

        usd_idx = None
        brent_idx = None
        date_idx = None

        for j, name in enumerate(header):
            if name is None:
                continue
            s = str(name).strip().lower()
            if usd_idx is None and ("usd" in s or "доллар" in s):
                usd_idx = j
            if brent_idx is None and ("brent" in s or "брент" in s):
                brent_idx = j
            if "дата" in s or "date" in s:
                date_idx = j

        target_row = None

        # Если есть колонка даты — выбираем строку с максимальной датой
        if date_idx is not None:
            latest_date = None

            for row in data_rows:
                if row is None or date_idx >= len(row):
                    continue
                cell_date = parse_excel_date(row[date_idx])
                if cell_date is None:
                    continue

                if latest_date is None or cell_date > latest_date:
                    latest_date = cell_date
                    target_row = row

            if target_row is None:
                # fallback: последняя непустая строка
                for row in reversed(data_rows):
                    if row and any(c is not None for c in row):
                        target_row = row
                        break
        else:
            # Нет колонки даты — берём последнюю непустую строку
            for row in reversed(data_rows):
                if row and any(c is not None for c in row):
                    target_row = row
                    break

        if not target_row:
            return {"usd_rate": None, "brent_price": None}

        def fmt(value):
            if value is None:
                return None
            if isinstance(value, (int, float)):
                s = f"{value:,.2f}"
                s = s.replace(",", " ").replace(".", ",")
                return s
            return str(value)

        if usd_idx is not None and usd_idx < len(target_row):
            usd_rate = fmt(target_row[usd_idx])

        if brent_idx is not None and brent_idx < len(target_row):
            brent_price = fmt(target_row[brent_idx])

    except Exception:
        usd_rate = None
        brent_price = None

    return {"usd_rate": usd_rate, "brent_price": brent_price}


def collect_all_dates() -> list[date]:
    """
    Собираем все даты из всех листов, где есть дата-колонка.
    Возвращаем уникальные даты по убыванию (последние – первые).
    """
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(
            f"Excel-файл не найден: {EXCEL_FILE}. "
            f"Проверь путь в config.py или имя файла."
        )

    wb = load_workbook(EXCEL_FILE, data_only=True)
    dates_set: set[date] = set()

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        header = rows[0]
        data_rows = rows[1:]

        # Ищем колонку с датой
        date_col_index = None
        for col_idx in range(len(header)):
            for row in data_rows:
                if row is None or col_idx >= len(row):
                    continue
                if parse_excel_date(row[col_idx]) is not None:
                    date_col_index = col_idx
                    break
            if date_col_index is not None:
                break

        if date_col_index is None:
            continue

        for row in data_rows:
            if row is None or date_col_index >= len(row):
                continue
            cell_date = parse_excel_date(row[date_col_index])
            if cell_date is not None:
                dates_set.add(cell_date)

    return sorted(dates_set, reverse=True)


def load_blocks_from_excel(date_filter: date | None = None) -> list[dict]:
    """
    Читает Excel и формирует список блоков (лист = блок).

    Если date_filter is None:
      - для листов с датой берём САМУЮ ПОСЛЕДНЮЮ дату на этом листе.
    Если date_filter задан:
      - для листов с датой берём только строки с этой датой (если она есть),
        иначе — ближайшую предыдущую дату на этом листе.
    Для листов без даты – берём все непустые строки.
    """
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(
            f"Excel-файл не найден: {EXCEL_FILE}. "
            f"Проверь путь в config.py или имя файла."
        )

    wb = load_workbook(EXCEL_FILE, data_only=True)
    blocks: list[dict] = []

    for ws in wb.worksheets:
        sheet_name = ws.title

        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            blocks.append(
                {
                    "id": sheet_name,
                    "title": sheet_name,
                    "sheetName": sheet_name,
                    "columns": [],
                    "rows": [],
                    "numericColumns": [],
                }
            )
            continue

        header = all_rows[0]
        columns = [str(c) if c is not None else "" for c in header]
        data_rows_raw = all_rows[1:]

        # ---------- Ищем дату-колонку ----------
        date_col_index = None
        date_to_rows: dict[date, list[tuple]] = {}

        for col_idx in range(len(columns)):
            for row in data_rows_raw:
                if row is None or col_idx >= len(row):
                    continue
                if parse_excel_date(row[col_idx]) is not None:
                    date_col_index = col_idx
                    break
            if date_col_index is not None:
                break

        target_date: date | None = None
        filtered_rows_raw = data_rows_raw

        if date_col_index is not None:
            # Собираем строки по датам
            for row in data_rows_raw:
                if row is None or date_col_index >= len(row):
                    continue
                d = parse_excel_date(row[date_col_index])
                if d is None:
                    continue
                date_to_rows.setdefault(d, []).append(row)

            if date_to_rows:
                if date_filter is not None:
                    if date_filter in date_to_rows:
                        target_date = date_filter
                    else:
                        # Берём ближайшую дату <= date_filter, либо последнюю вообще
                        earlier = [d for d in date_to_rows.keys() if d <= date_filter]
                        if earlier:
                            target_date = max(earlier)
                        else:
                            target_date = max(date_to_rows.keys())
                else:
                    target_date = max(date_to_rows.keys())

                filtered_rows_raw = date_to_rows.get(target_date, [])

        # ---------- Предыдущий день для листа "Конкуренты" ----------
        prev_date: date | None = None
        prev_values: dict[str, dict[str, float]] | None = None

        if (
            sheet_name == "Конкуренты"
            and date_col_index is not None
            and date_to_rows
            and target_date is not None
        ):
            earlier_dates = [d for d in date_to_rows.keys() if d < target_date]
            if earlier_dates:
                prev_date = max(earlier_dates)
                prev_rows_raw = date_to_rows.get(prev_date, [])

                # Определяем колонку "Продукт"
                product_col_index = 0
                for idx, col_name in enumerate(columns):
                    s = str(col_name or "").strip().lower()
                    if "продукт" in s or "номенклат" in s:
                        product_col_index = idx
                        break

                prev_values = {}

                for row in prev_rows_raw:
                    if row is None or product_col_index >= len(row):
                        continue
                    product_cell = row[product_col_index]
                    if product_cell is None:
                        continue
                    product_key = str(product_cell).strip()
                    if not product_key:
                        continue

                    row_map = prev_values.setdefault(product_key, {})

                    for j, col_name in enumerate(columns):
                        if j >= len(row):
                            continue
                        val = row[j]
                        if isinstance(val, (int, float)):
                            row_map[str(col_name) if col_name is not None else ""] = val

        # ---------- Преобразование строк для фронта ----------
        rows: list[list] = []
        numeric_flags = [False] * len(columns)

        for row_raw in filtered_rows_raw:
            if row_raw is None:
                continue

            raw_list = list(row_raw[: len(columns)])
            if all(cell is None for cell in raw_list):
                continue

            display_row: list = []

            for idx, cell in enumerate(raw_list):
                if isinstance(cell, (int, float)):
                    numeric_flags[idx] = True

                cell_date_val = parse_excel_date(cell)
                if cell_date_val is not None:
                    display_row.append(cell_date_val.strftime("%Y-%m-%d"))
                else:
                    display_row.append(cell)

            rows.append(display_row)

        numeric_indices = [idx for idx, is_num in enumerate(numeric_flags) if is_num]

        block: dict = {
            "id": sheet_name,
            "title": sheet_name,
            "sheetName": sheet_name,
            "columns": columns,
            "rows": rows,
            "numericColumns": numeric_indices,
        }

        # Для "Конкурентов" дополнительно отправляем вчерашнюю дату и значения
        if sheet_name == "Конкуренты":
            block["prevDate"] = prev_date.isoformat() if prev_date else None
            block["prevValues"] = prev_values or {}

        blocks.append(block)

    return blocks


# ---------------------------- ROUTES ---------------------------------


@app.route("/")
def index():
    """
    Главная страница: только рендер HTML.
    Курсы подставляем сразу, дату обновления и архив — через JS (/api/dates).
    """
    metrics = get_header_metrics()
    return render_template("index.html", **metrics)


@app.route("/api/blocks")
def api_blocks():
    """
    API: отдать блоки.
    Параметр ?date=YYYY-MM-DD — дата, за которую нужно отдать блоки.
    Если не передавать ?date, вернём "последний день" по каждому листу.
    """
    try:
        date_param = request.args.get("date")
        target_date: date | None = None

        if date_param:
            target_date = parse_excel_date(date_param)
            if target_date is None:
                return (
                    jsonify({"error": f"Некорректный формат даты: {date_param}"}),
                    400,
                )

        blocks = load_blocks_from_excel(date_filter=target_date)
        return jsonify({"blocks": blocks})
    except FileNotFoundError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        return jsonify({"error": f"Ошибка при чтении Excel: {e}"}), 500


@app.route("/api/dates")
def api_dates():
    """
    API: список всех доступных дат (для архива).
    Возвращаем даты в формате YYYY-MM-DD по убыванию.
    """
    try:
        dates = collect_all_dates()
        dates_str = [d.isoformat() for d in dates]
        return jsonify({"dates": dates_str})
    except FileNotFoundError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        return jsonify({"error": f"Ошибка при чтении дат из Excel: {e}"}), 500


@app.route("/api/screenshot", methods=["POST"])
def api_screenshot():
    """
    Принимает base64 PNG из фронта и сохраняет его в папку screenshots
    рядом с Excel-файлом (в структуре проекта).
    """
    try:
        payload = request.get_json()
        if not payload or "imageData" not in payload:
            return jsonify({"error": "Не переданы данные изображения"}), 400

        image_data = payload["imageData"]
        if not isinstance(image_data, str) or "," not in image_data:
            return jsonify({"error": "Неверный формат данных изображения"}), 400

        header, b64data = image_data.split(",", 1)
        if "base64" not in header:
            return jsonify({"error": "Ожидался base64-формат изображения"}), 400

        img_bytes = base64.b64decode(b64data)

        screenshots_dir = EXCEL_FILE.parent / "screenshots"
        screenshots_dir.mkdir(parents=True, exist_ok=True)

        now = datetime.now()
        filename = f"trastboard_{now.strftime('%Y-%m-%d_%H-%M-%S')}.png"
        filepath = screenshots_dir / filename

        with open(filepath, "wb") as f:
            f.write(img_bytes)

        return jsonify({"status": "ok", "fileName": filename})
    except Exception as e:
        return jsonify({"error": f"Ошибка сохранения скрина: {e}"}), 500


if __name__ == "__main__":
    app.run(debug=True)
